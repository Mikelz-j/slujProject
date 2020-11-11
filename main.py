import openpyxl
from time import time

tic = time()

old_pages = "sat_list_solutions.xlsx"
id_pages = ""

wb = openpyxl.reader.excel.load_workbook(filename=old_pages)
ws = wb.active

wb_id = openpyxl.reader.excel.load_workbook(filename=id_pages)
ws_id = wb_id.active
items = {}
dd = {}
for i in range(2,48):
    key_id = str(ws_id['B' + str(i)].value)
    val_id = str(ws_id['A' + str(i)].value)
    dd[key_id] =val_id

def find_id_item(items):
    list_id = []
    for i in items:
        s = i.replace(' ', '')
        list_id.append(s[-6:-1])
    return list_id

def replace_id(items):
    list_replace_id = []
    for i in items:
        s = i.replace(' ', '')
        if s[-6:-1] in dd:
            rez_str = i.replace(s[-6:-1], dd[s[-6:-1]])
        else:
            rez_str = i.replace(s[-6:-1], 'error')
        list_replace_id.append(rez_str)
    return list_replace_id

for i in range(2,135):
    key = ws['A' + str(i)].value
    if ws['B' + str(i)].value:
        val = (ws['B' + str(i)].value).split('/')
    else:
        val = []
    items[key] =replace_id(val)


r = open("rez.txt", "a")
for i in items:
    r.write(f"{i} - {find_id_item(items[i])}" + '\n' )
r.close()

print('Ok')
toc = time()
print(str(round((toc - tic), 1)) + ' sec')